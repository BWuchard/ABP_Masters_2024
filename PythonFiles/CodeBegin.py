import pandas as pd
import openpyxl
import re
import numpy as np
import os
from collections import defaultdict


# Preliminary Information
year = 2021
antigens = ['DT', 'DTP', 'HepB', 'Hib', 'HPV', 'MMR', 'PCV', 'Rota', 'TCV', 'YF', 'OPV']
antigen_doses = {'DT':3,'DTP':3,'HepB':3,'Hib':3,'HPV':2,'JE':2, 'MenA':3,'MMR':2,'OPV':4,'PCV':3,'Rota':2,'TCV':3,'YF':1}
print(antigens)
n_antigens = len(antigens)
print(n_antigens)
vaccines = ['DT', 'Hib', 'DTP', 'DTP-HepB-Hib-IPV', 'DTP-HepB-IPV', 'DTP-Hib', 'DTP-Hib-IPV', 'DTP-IPV', 'DTP', 'DTP-HepB-Hib', 'HepB', 'HPV', 'IPV', 'MMR', 'PCV', 'Rota', 'TCV', 'YF', 'OPV'] 

# Experimental Lists
PAHO = ['Brazil', 'Mexico', 'Bahamas', 'Barbados', 'Trinidad', 'Chile', 'Uruguay', 'St. Kitts and Nevis', 'Antigua', 'Venezuela', 'Argentina', 'Panama', 'Suriname', 'Costa Rica', 'Grenada', 'Colombia', 'Saint Lucia', 'Dominica', 'Saint Vincent', 'Cuba', 'Peru', 'Dominican Rep.', 'Ecuador', 'Jamaica', 'Belize', 'El Salvador', 'Guyana', 'Paraguay', 'Guatemala', 'Bolivia', 'Honduras', 'Nicaragua', 'Haiti']
PAHOUSC = ['Brazil', 'Mexico', 'Bahamas', 'Barbados', 'Trinidad', 'Chile', 'Uruguay', 'St. Kitts and Nevis', 'Antigua', 'Venezuela', 'Argentina', 'Panama', 'Suriname', 'Costa Rica', 'Grenada', 'Colombia', 'Saint Lucia', 'Dominica', 'Saint Vincent', 'Cuba', 'Peru', 'Dominican Rep.', 'Ecuador', 'Jamaica', 'Belize', 'El Salvador', 'Guyana', 'Paraguay', 'Guatemala', 'Bolivia', 'Honduras', 'Nicaragua', 'Haiti', 'USA', 'Canada']
Gavi = ["Afghanistan", "Benin", "Burkina Faso", "Burundi", "Cambodia", "CAR", "Chad", "Comoros", "DRC", "Eritrea", "Ethiopia", "Gambia", "Guinea", "Guinea-Bissau", "Haiti", "Kyrgyzstan", "Lesotho", "Liberia", "Madagascar", "Malawi", "Mali", "Mauritania", "Mozambique", "Myanmar", "Nepal", "Niger", "North Korea", "Rwanda", "Senegal", "Sierra Leone", "Somalia", "South Sudan", "Sudan", "Syria", "Tajikistan", "Tanzania", "Togo", "Uganda", "Yemen", "Zambia", "Zimbabwe", "Bangladesh", "Cameroon", "Congo", "Côte d'Ivoire", "Djibouti", "Ghana", "Kenya", "Laos", "Nigeria", "Pakistan", "Papua New Guinea", "São Tomé", "Solomon Isl."]
#############################################################################################

####
#### Organizing and conjoining of raw data. Column, vaccine, and company names are changed to be uniform. Years are filtered to the year set in the preliminary information #####
####

############################################################################################

country_status_df = pd.read_csv('C:\\Users\\Bailey\\Downloads\\Research\\11_2021_Country_Status_(Server)_data.csv')

# Create a regular expression pattern by joining the filter_list with '|'
pattern = '|'.join(antigens)

# CDC Price data
cdc_pr_df = pd.read_csv('C:\\Users\\Bailey\\Downloads\\Research\\98_CDC_Pricing_data.csv')
cdc_pr_df['Vaccine '] = cdc_pr_df['Vaccine '].replace({'MenACYW-135 Ps':'MenA','MenACW-135 Ps':'MenA','MenAC Ps':'MenA','JE live att.':'JE', 'TT':'Td','DTaP':'DTP','DTaP-HepB-IPV':'DTP-HepB-IPV','DTaP-HepB-Hib-IPV':'DTP-HepB-Hib-IPV','DTaP-HepB-IPV':'DTP-HepB-IPV','DTaP-Hib-IPV':'DTP-Hib-IPV','DTaP-IPV':'DTP-IPV','HepA+B':'HepB','HPV2':'HPV','HPV4':'HPV','HPV9':'HPV','MMRV':'MMR','PCV13':'PCV','Tdap':'Td','bOPV':'OPV','DTwP':'DTP','DTwP-HepB-Hib':'DTP-HepB-Hib','DTwP-Hib':'DTP-Hib','PCV10':'PCV','tOPV':'OPV'})
cdc_pr_df = cdc_pr_df[cdc_pr_df['Year'] == year]
cdc_pr_df.rename(columns={'Vaccine ': 'Vaccine'}, inplace=True)
cdc_pr_df = cdc_pr_df[cdc_pr_df['Vaccine'].str.contains(pattern)]
cdc_pr_df = cdc_pr_df[~cdc_pr_df['Vaccine'].str.contains('adult')]
cdc_pr_df = cdc_pr_df[~cdc_pr_df['Vaccine'].str.contains(r'\badult\b', regex=True)]
cdc_pr_df.rename(columns={'Company ': 'Company'}, inplace=True)
cdc_pr_df['Company'] = cdc_pr_df['Company'].replace({'AstraZen.':'AstraZeneca'})
cdc_pr_df = cdc_pr_df.dropna(subset=['Avg. Price'])

# Paho price data
paho_pr_df = pd.read_csv('C:\\Users\\Bailey\\Downloads\\Research\\99_PAHO_Pricing_data.csv')
paho_pr_df['Vaccine '] = paho_pr_df['Vaccine '].replace({'MenACYW-135 Ps':'MenA','MenACW-135 Ps':'MenA','MenAC Ps':'MenA','JE live att.':'JE', 'TT':'Td','DTaP':'DTP','DTaP-HepB-IPV':'DTP-HepB-IPV','DTaP-HepB-Hib-IPV':'DTP-HepB-Hib-IPV','DTaP-HepB-IPV':'DTP-HepB-IPV','DTaP-Hib-IPV':'DTP-Hib-IPV','DTaP-IPV':'DTP-IPV','HepA+B':'HepB','HPV2':'HPV','HPV4':'HPV','HPV9':'HPV','MMRV':'MMR','PCV13':'PCV','Tdap':'Td','bOPV':'OPV','DTwP':'DTP','DTwP-HepB-Hib':'DTP-HepB-Hib','DTwP-Hib':'DTP-Hib','PCV10':'PCV','tOPV':'OPV'})
paho_pr_df = paho_pr_df[paho_pr_df['Year'] == year]
paho_pr_df.rename(columns={'Vaccine ': 'Vaccine'}, inplace=True)
paho_pr_df = paho_pr_df[paho_pr_df['Vaccine'].str.contains(pattern)]
paho_pr_df = paho_pr_df[~paho_pr_df['Vaccine'].str.contains('adult')]
paho_pr_df = paho_pr_df[~paho_pr_df['Vaccine'].str.contains(r'\badult\b', regex=True)]
paho_pr_df.sort_values('Vaccine', inplace=True)
paho_pr_df['Avg. Price'].fillna(method='ffill', inplace=True)
paho_pr_df['Avg. Price'].fillna(method='bfill', inplace=True)
paho_pr_df = paho_pr_df.dropna(subset=['Avg. Price'])
paho_pr_df.rename(columns={'Country of Origin':'Country'}, inplace=True)

# UNICEF Price data
unicef_pr_df = pd.read_csv('C:\\Users\\Bailey\\Downloads\\Research\\910_UNICEF_Pricing_data.csv')
unicef_pr_df = unicef_pr_df.dropna(subset=['Avg. Price'])
unicef_pr_df['Vaccine '] = unicef_pr_df['Vaccine '].replace({'MenACYW-135 Ps':'MenA','MenACW-135 Ps':'MenA','MenAC Ps':'MenA','JE live att.':'JE', 'TT':'Td','DTaP':'DTP','DTaP-HepB-IPV':'DTP-HepB-IPV','DTaP-HepB-Hib-IPV':'DTP-HepB-Hib-IPV','DTaP-HepB-IPV':'DTP-HepB-IPV','DTaP-Hib-IPV':'DTP-Hib-IPV','DTaP-IPV':'DTP-IPV','HepA+B':'HepB','HPV2':'HPV','HPV4':'HPV','HPV9':'HPV','MMRV':'MMR','PCV13':'PCV','Tdap':'Td','bOPV':'OPV','DTwP':'DTP','DTwP-HepB-Hib':'DTP-HepB-Hib','DTwP-Hib':'DTP-Hib','PCV10':'PCV','tOPV':'OPV'})
unicef_pr_df = unicef_pr_df[unicef_pr_df['Year'] == year]
unicef_pr_df.rename(columns={'Vaccine ': 'Vaccine'}, inplace=True)
unicef_pr_df = unicef_pr_df[unicef_pr_df['Vaccine'].str.contains(pattern)]
unicef_pr_df = unicef_pr_df[~unicef_pr_df['Vaccine'].str.contains('adult')]
unicef_pr_df = unicef_pr_df[~unicef_pr_df['Vaccine'].str.contains(r'\badult\b', regex=True)]
unicef_pr_df.sort_values('Vaccine', inplace=True)
unicef_pr_df.rename(columns={'Company  ': 'Company'}, inplace=True)
unicef_pr_df['Company'] = unicef_pr_df['Company'].replace({'BBIL':'Bharat Biotech International Limited','BioE':'Biological E. Limited','Biological E':'Biological E. Limited','BioM':'Bio-Med','CNBG Chengdu':'CNBG / Chengdu','GreenSignal':'Green SB','Haffkine':'Haffkine Bio-Pharmaceutical Corp','InterVax':'Intervax','Japan BCG Lab':'Japan BCG Laboratory','LG Chem':'LG Life Science','Panacea':'Panacea Biotec','SII':'Serum Institute Of India Ltd.'})
unicef_pr_df['Avg. Price'].fillna(method='ffill', inplace=True)
unicef_pr_df['Avg. Price'].fillna(method='bfill', inplace=True)
unicef_pr_df = unicef_pr_df.dropna(subset=['Avg. Price'])

# Capacity data, to be changed because it currently has a confidential file
capacity_df = pd.read_excel('C:\\Users\\Bailey\\Downloads\\Research\\Capacity and Availability from Market Insights CONFIDENTIAL.xlsx')
capacity_df['Vaccine'] = capacity_df['Vaccine'].replace({'Diphtheria-Tetanus':'DT', 'Diphtheria-Tetanus-Pertussis (whole cell)':'DTP','Diphtheria-Tetanus-Pertussis (whole cell)-Hepatitis B-Haemophilus influenzae type b':'DTP-HepB-Hib', 'Human Papilloma Virus (HPV)':'HPV','IPV-Salk':'IPV','Measles, Mumps and Rubella':'MMR','Polio Vaccine - Inactivated (IPV)':'IPV','Polio Vaccine - Oral':'bOPV', 'Rotavirus':'Rota', 'Typhoid':'Td', 'tOPV, bOPV type 1 & 3': 'bOPV', 'tOPV, bOPV type 1 & 3, mOPV type 1/3':'bOPV', 'Tetanus Toxoid':'Td','Pneumococcal (conjugate)':'PCV','Japanese Encephalitis':'YF'})
capacity_df = capacity_df[capacity_df['Vaccine'].isin(antigens)]
capacity_df.rename(columns={'Company Name': 'Company'}, inplace=True)
capacity_df = capacity_df[capacity_df['Year'] <= 2021]
capacity_df = capacity_df.sort_values(by='Year', ascending=False)
capacity_df = capacity_df.drop_duplicates(subset=['Vaccine','Company'])
capacity_df.rename(columns={'Country Name':'Country'}, inplace=True)

# Average pricing data
average_pr_df = pd.read_csv('C:\\Users\\Bailey\\Downloads\\Research\\42_Average_Vaccine_Pricing_Reference_Table_data.csv')

# C Calculating data
c_calculating_df = pd.read_excel('C:\\Users\\Bailey\\Downloads\\Research\\CNumbers.xlsx')

# Creating a set of all companies present in the dataframes
companies = set()

for df in [cdc_pr_df, unicef_pr_df, capacity_df]:
    companies.update(df['Company'].tolist())

print(companies)

company_dict = {}

for df in [cdc_pr_df, unicef_pr_df, capacity_df]:
    for _, row in df.iterrows():
        company = row['Company']
        vaccine = row['Vaccine']
        if company not in company_dict:
            company_dict[company] = set()
        company_dict[company].add(vaccine)

# Print the resulting dictionary - can be taken out later
for company, vaccines in company_dict.items():
    print(f"{company}: {', '.join(vaccines)}")

# Identify companies in cdc_pr_df and unicef_pr_df that are not in capacity_df
companies_to_add = set(cdc_pr_df['Company']).union(set(unicef_pr_df['Company'])) - set(capacity_df['Company'])
#Identify companies that are common between cdc_pr_df, unicef_pr_df and capacity_df
common_companies = set(cdc_pr_df['Company']).intersection(set(unicef_pr_df['Company']), set(capacity_df['Company']))

# Create a new dataframe with the companies to add
new_rows = []
for company in companies_to_add:
    vaccines = company_dict.get(company, set())
    for vaccine in vaccines:
        vaccine_capacity_mean = capacity_df.loc[capacity_df['Vaccine'] == vaccine, 'Capacity, Doses (Total)'].mean()
        new_rows.append({'Company': company, 'Vaccine': vaccine, 'Capacity, Doses (Total)': vaccine_capacity_mean})

# Append the new rows to capacity_df
df_to_add = pd.DataFrame(new_rows)
capacity_df = pd.concat([capacity_df, df_to_add], ignore_index=True)

#Create a dataframe for the common companies
for company in common_companies:
    vaccines_cdc_pr_df_unicef_pr_df = company_dict.get(company, set())
    vaccines_capacity_df = set(capacity_df[capacity_df['Company'] == company]['Vaccine'])
    additional_vaccines = vaccines_cdc_pr_df_unicef_pr_df - vaccines_capacity_df
    for vaccine in additional_vaccines:
        new_rows.append({'Company': company, 'Vaccine': vaccine})
df_to_add = pd.DataFrame(new_rows)
# Append the new dataframe to capacity_df
capacity_df = pd.concat([capacity_df, df_to_add], ignore_index=True)

#Generate data for the capacity of the common companies that didn't have existing information. Estimated to be an average of the existing data for vaccines that had data, populated for the full company
for company in common_companies:
    company_vaccines = set(capacity_df[capacity_df['Company'] == company]['Vaccine'])
    company_capacity_mean = capacity_df.loc[(capacity_df['Company'] == company) & capacity_df['Vaccine'].isin(company_vaccines), 'Capacity, Doses (Total)'].mean()
    capacity_df.loc[(capacity_df['Company'] == company) & capacity_df['Vaccine'].isin(company_vaccines) & capacity_df['Capacity, Doses (Total)'].isnull(), 'Capacity, Doses (Total)'] = company_capacity_mean

# Calculate the overall average capacity for non-empty values in capacity_df
overall_capacity_mean = capacity_df.loc[capacity_df['Capacity, Doses (Total)'].notnull(), 'Capacity, Doses (Total)'].mean()

# Fill the remaining empty values with the overall average capacity
capacity_df['Capacity, Doses (Total)'].fillna(overall_capacity_mean, inplace=True)

# Fill the empty values in 'Year' column with the default year
capacity_df['Year'].fillna(year, inplace=True)

def populate_company(row):
    if pd.isnull(row['Country']):
        same_vaccine_rows = capacity_df[capacity_df['Vaccine'] == row['Vaccine']]
        random_company = np.random.choice(same_vaccine_rows['Company'])
        return random_company
    else:
        matching_entry = capacity_df[(capacity_df['Vaccine'] == row['Vaccine']) & (capacity_df['Country'] == row['Country'])]
        if not matching_entry.empty:
            return matching_entry.iloc[0]['Company']
        else:
            vaccine_companies = capacity_df[capacity_df['Vaccine'] == row['Vaccine']]['Company']
            if not vaccine_companies.empty:
                majority_company = vaccine_companies.mode().values[0]
                return majority_company
            else:
                return np.random.choice(list(companies))

# Apply the function to populate the Company column in paho_pr_df
paho_pr_df['Company'] = paho_pr_df.apply(populate_company, axis=1)


# Output data to excel sheet to test data
csv_writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')
country_status_df.to_excel(csv_writer, sheet_name='Country', index = False)
cdc_pr_df.to_excel(csv_writer, sheet_name='CDC_Price', index = False)
paho_pr_df.to_excel(csv_writer, sheet_name='PAHO_Price', index = False)
unicef_pr_df.to_excel(csv_writer, sheet_name='UNICEF_Price', index = False)
capacity_df.to_excel(csv_writer, sheet_name='Capacity', index = False)
average_pr_df.to_excel(csv_writer, sheet_name='Average_Price', index = False)
c_calculating_df.to_excel(csv_writer, sheet_name='CNumbers', index = False)

csv_writer.save()

csv_file = pd.ExcelFile('output.xlsx')

# Get the sheet names
sheet_names = csv_file.sheet_names

# Create an empty dictionary to store the dataframes
dataframes = {}

# Read each sheet into a dataframe
for sheet_name in sheet_names:
    dataframes[sheet_name] = pd.read_excel(csv_file, sheet_name)

country_status_df = dataframes['Country']
cdc_pr_df = dataframes['CDC_Price']
paho_pr_df = dataframes['PAHO_Price']
unicef_pr_df = dataframes['UNICEF_Price']
capacity_df = dataframes['Capacity']
average_pr_df = dataframes['Average_Price']
c_calculating_df = dataframes['CNumbers']


####################################################################################################################################################################

### Select Countries ###

#########################################################################################

file_path = 'C:\\Users\\Bailey\\Downloads\\Research\\output.xlsx'

selected_list = Gavi


ncountries = len(selected_list) # parameter for n_country
print(ncountries)

# Country parameter for data file, needs a code for each country
country_param_df = country_status_df[country_status_df['Country Name'].isin(selected_list)]
country_param_df = country_param_df.drop(columns=['WHO Region','Current WB Status','Gavi Eligibility', 'Gavi Co-financing'])
country_param_df.insert(loc=3, column='target_country', value=0)
country_param_df = country_param_df.rename(columns={'Country Name' : 'country_name', 'Birth Cohort ' : 'annual_births', 'GDP per capita' : 'GNI_c'})
country_param_df = country_param_df.sort_values(by='country_name')
country_param_df = country_param_df.reset_index(drop=True)
country_param_df.index = country_param_df.index + 1
print(country_param_df)

excel_file = openpyxl.load_workbook(file_path)
sheet_name = "CountriesSelected"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    country_param_df.to_excel(writer, sheet_name=sheet_name, index=False)

#########################################################################################

### Reservation Price Weights ###

#########################################################################################

cp_df_sorted = pd.read_excel('PAHOcriteria.xlsx')
cp_df_sorted = cp_df_sorted.rename(columns={'country_name' : 'Country', 'Birth Cohort' : 'Annual_births', 'GNI_c' : 'GNI_pc'})
cp_df_sorted = cp_df_sorted.sort_values(by='Country')
cp_df_sorted = cp_df_sorted.reset_index(drop=True)
cp_df_sorted['CCode'] = cp_df_sorted.index + 1
cp_df_sorted = cp_df_sorted.replace(' ','_', regex=True)
print(cp_df_sorted)

column_name = 'CCode'
columns = cp_df_sorted.columns.tolist()
columns.insert(0, columns.pop(columns.index(column_name)))
cp_df_sorted = cp_df_sorted.reindex(columns=columns)

excel_file = openpyxl.load_workbook(file_path)
sheet_name = "CNTY"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    cp_df_sorted.to_excel(writer, sheet_name=sheet_name, index=False)

cnty = cp_df_sorted
#Create an empty matrix with a numbe of rows and columns equal to the number of countries to consider
M=np.zeros([len(cnty),len(cnty)], dtype=float)
    
#Populate the matrix with the proxy of their relative weights (gni_pc for this implementation) 
#Any other proxy can be used. Notice that the ranges go from 1-num of coutries, yet matrix is index from 0 to num_cty -1
for i in range(len(cnty)):
    for j in range(len(cnty)):
        M[i, j] = cnty['GNI_pc'].iloc[i] / cnty['GNI_pc'].iloc[j]

#Compute eigen vectors and eigen values for the proxies of the relative country importances
w,v=np.linalg.eig(M) 
    
#Create an array with the indices of where the eigen values that are significant and have no significant
#imaginary parts. Asign to the array the position of the first eigenvalue that meets these characteristics
princ=[]
princ=np.where((abs(w.real)>=0.0000001)&(abs(w.imag)<0.00000001))[0]
    
ind_interest=-1
    
#If there are real principal eigen vectors proceed with the example
if len(princ)>0: 
    ind_interest=princ[0]
else:
    print("Warning: No real eigen vector exists")
    
# Get the countries and their corresponding relative weights
countries = [cnty['Country'][i] for i in range(len(cnty))]
relative_weights = v[:, ind_interest].real / np.average(v[:, ind_interest].real)

# Create ResPriceFrame with countries and relative weights
ResPriceFrame = pd.DataFrame({'Country': countries, 'Relative_weight': relative_weights})

#    Frame.to_csv("C:\Users\Bailey\Downloads\Code_Mkt\RCalculation\FrameRP.csv")

excel_file = openpyxl.load_workbook(file_path)
sheet_name = "res_pric_weights"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    ResPriceFrame.to_excel(writer, sheet_name=sheet_name, index=False)

###############################################################
       
       ### Market Code ###
       
###############################################################

# Read the Excel file
excel_file = 'C:\\Users\\Bailey\\Downloads\\Research\\GaviPAHORecreation\\PAHO\\2M_g_b_bh.xlsx'  # Replace with the actual file path
df = pd.read_excel(excel_file, header=None)

# Extract the market assignments from the DataFrame
num_markets = len(df.index)

# Create empty market sets
markets = {}
for i in range(1, num_markets + 1):
    markets[i] = set()

# Create empty dictionary for country codes
country_codes = {}

# Function to generate country codes
def generate_country_codes(country_list):
    sorted_countries = sorted(country_list)  # Sort the country list alphabetically
    for index, country in enumerate(sorted_countries, start=1):
        country_codes[country] = index

# Extract country names and market assignments from the Excel file
country_names = []
for row in df.iterrows():
    market_num = row[1][0]  # Get the market number from the first column
    countries = row[1][1:].dropna().tolist()  # Get the list of countries in the market
    markets[market_num].update(countries)  # Assign countries to the market
    country_names.extend(countries)


# Generate country codes
generate_country_codes(country_names)

marketa = pd.DataFrame(columns=["Country"] + [str(i) for i in range(1, num_markets + 1)])

used_countries = set()

for market in range(1, num_markets + 1):
    print("\nMarket", market)
    print("Assigned countries:", ", ".join(c for c in markets[market]))

    for country in markets[market]:
        markets[market].add(country)
        used_countries.add(country)
        country_names.remove(country)
        marketa.loc[len(marketa)] = [country_codes[country]] + [1 if country in markets[i] else 0 for i in range(1, num_markets + 1)]

print("\nMarket assignments:")
print(marketa)

# Sort the remaining countries in alphabetical order
remaining_countries = sorted(country_names)

print("\nRemaining countries:")
for country in remaining_countries:
    print(f"{country_codes[country]}: {country}")

# Create a new dataframe to store the market averages
market_averages = pd.DataFrame(columns=['Market', 'Country', 'gni_p', 'l'])

# Iterate over each market and its assigned countries
for market, selected_user in markets.items():
    # Filter the original dataframe for countries in the current market
    market_data = country_param_df[country_param_df['country_name'].isin(selected_user)]

    # Calculate the average GNI per capita and annual births for the market
    average_gni = market_data['GNI_c'].mean()
    average_births = market_data['annual_births'].sum()

    # Create a new row in the market_averages dataframe
    country_codes_str = ', '.join([str(country_codes[country]) for country in selected_user])
    market_averages = market_averages.append({'Market': market, 'Country': country_codes_str,
                                              'gni_p': average_gni, 'l': average_births},
                                             ignore_index=True)

print(market_averages)

excel_file = openpyxl.load_workbook(file_path)
sheet_name = "market_averages"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    market_averages.to_excel(writer, sheet_name=sheet_name, index=False)



########################################################################################

    ### Antigens, Bundles, Manufacturers ###

########################################################################################

sets = {'PAHO', 'PAHOUSC'}

if selected_list is PAHO:
    selected_dataframe = paho_pr_df
if selected_list is Gavi:
    selected_dataframe = unicef_pr_df

### Set A ###
# Extract unique antigens from the 'Vaccine' column, overwrite the initial antigen set
antigens = set()
for vaccine in selected_dataframe['Vaccine']:
    vaccines = vaccine.split('-')
    for v in vaccines:
        antigens.add(v.strip())
sorted_antigens = sorted(antigens)

# Create a dictionary to store the mapping of antigens to codes
antigen_codes = {antigen: code for code, antigen in enumerate(sorted_antigens, start=1)}
print(antigen_codes)
           
# Set to store producers
setp = set(selected_dataframe['Company'])
producers_as_strings = [str(p) for p in setp]
sorted_setp = sorted(producers_as_strings)

paramk = {producer: code for code, producer in enumerate(sorted_setp, start=1)}

print(paramk)

################################################################################################

### Bundles and A1, B1, B2 Parameter ###

################################################################################################

# Create a list to store the vaccine bundles
vaccine_bundles = []

# Iterate over each vaccine in the "Vaccine" column
for vaccine in selected_dataframe['Vaccine']:
    vaccine_bundles.append(vaccine)

# Print the vaccine bundles along with their codes and antigen values
for index, bundle in enumerate(vaccine_bundles, start=1):
    antigens = []  # List to store the antigen codes for the current bundle
    
    # Extract the antigens from the bundle and map them to their codes
    for antigen in bundle.split('-'):
        if antigen in antigen_codes:
            antigens.append(str(antigen_codes[antigen]))
    
    print(f"A1[{index}]:= {' '.join(antigens)} ;")

    

##########
### B1 ###
##########
    
# Create a dictionary to map antigen codes to bundle indices
antigen_bundle_indices = {}

# Iterate over each vaccine in the "Vaccine" column
for index, vaccine in enumerate(selected_dataframe['Vaccine'], start=1):

    # Extract the antigens from the bundle and check if they are in the antigen codes
    for antigen in vaccine.split('-'):
        if antigen in antigen_codes:
            antigen_code = antigen_codes[antigen]
            if antigen_code in antigen_bundle_indices:
                antigen_bundle_indices[antigen_code].append(index)
            else:
                antigen_bundle_indices[antigen_code] = [index]

# Print the bundle indices that contain each antigen
for antigen_code, bundle_indices in antigen_bundle_indices.items():
    antigen = next(key for key, value in antigen_codes.items() if value == antigen_code)
    print(f"B1[{antigen_code}] := {' '.join(map(str, bundle_indices))} ;")

##########
### B2 ###
##########

# Create a dictionary to map producer codes to vaccine indices
producer_vaccine_indices = {}

# Create a dictionary to map vaccine indices to vaccine names
vaccine_names = {}

# Iterate over each vaccine index and its corresponding producer
for index, producer in enumerate(selected_dataframe['Company'], start=1):
    if producer in paramk:
        producer_code = paramk[producer]
        if producer_code in producer_vaccine_indices:
            producer_vaccine_indices[producer_code].append(index)
        else:
            producer_vaccine_indices[producer_code] = [index]
        
        # Add the vaccine name to the vaccine_names dictionary
        vaccine_name = selected_dataframe.loc[index - 1, 'Vaccine']
        vaccine_names[index] = vaccine_name

# Print the vaccine indices associated with each producer
for producer_code, vaccine_indices in producer_vaccine_indices.items():
    producer = next(key for key, value in paramk.items() if value == producer_code)
    print(f"B2[{producer_code}] := {' '.join(map(str, vaccine_indices))} ;")

# Print the vaccine names associated with each vaccine index
for index, vaccine_name in vaccine_names.items():
    print(f"Vaccine Name[{index}] := {vaccine_name}")
    
#############################################################################
    
    ### S k name C ###
    
#############################################################################



country_param_df_sum = pd.read_excel('PAHOcriteria.xlsx')
country_param_df_sum = country_param_df_sum['Birth Cohort'].sum()
country_status_df_sum = pd.read_csv('birthprojections.csv')
country_status_df_sum = country_status_df_sum[country_status_df_sum['Year'] == year]
country_status_df_sum = country_status_df_sum['Births - Sex: all - Age: all - Variant: estimates'].sum()


precresult = country_param_df_sum / country_status_df_sum

# Create a new DataFrame to store the new entries
new_entries = []

# Iterate over each producer code and its corresponding vaccine indices
for producer_code, vaccine_indices in producer_vaccine_indices.items():
    producer = next(key for key, value in paramk.items() if value == producer_code)
    
    # Iterate over each vaccine index for the current producer
    for index in vaccine_indices:
        vaccine_name = vaccine_names.get(index, None)
        if vaccine_name:
            # Create a new entry with the vaccine name, None for the investment, and the producer code
            new_entry = {'Vaccine': vaccine_name, 'Investment': None, 'Company': producer_code}
            new_entries.append(new_entry)

# Append the new entries to the c_calculating_df DataFrame
c_calculating_df = c_calculating_df.append(new_entries, ignore_index=True)

# Replace the entries in the 'Vaccine' column with the corresponding vaccine names
c_calculating_df['Vaccine'] = c_calculating_df['Vaccine'].map(lambda x: vaccine_names.get(x, x))

# Create a dictionary to map producer codes to producer names
producer_names = {code: producer for producer, code in paramk.items()}

# Replace producer codes with producer names in c_calculating_df
c_calculating_df['Company'] = c_calculating_df['Company'].map(producer_names)

# Create a dictionary mapping (vaccine, company) to capacity values from capacity_df
capacity_map = capacity_df.set_index(['Vaccine', 'Company'])['Capacity, Doses (Total)'].to_dict()

# Update the 'Capacity, Doses (Total)' column in c_calculating_df
c_calculating_df['Capacity, Doses (Total)'] = c_calculating_df.apply(lambda row: capacity_map.get((row['Vaccine'], row['Company']), None), axis=1)

# Calculate the overall average capacity for non-empty values in c_calculating_df
overall_capacity_mean = c_calculating_df.loc[c_calculating_df['Capacity, Doses (Total)'].notnull(), 'Capacity, Doses (Total)'].mean()

# Fill the remaining empty values in 'Capacity, Doses (Total)' column with the overall average capacity
c_calculating_df['Capacity, Doses (Total)'].fillna(overall_capacity_mean, inplace=True)

# Group c_calculating_df by the 'Vaccine' column and extract the 'Investment' column as a Series
investment_series = c_calculating_df.groupby('Vaccine')['Investment'].first()

# Map the investment values to the matching vaccines in c_calculating_df, but only for NaN values
c_calculating_df['Investment'] = c_calculating_df['Investment'].fillna(c_calculating_df['Vaccine'].map(investment_series))

# Identify combination vaccines
combination_vaccines = c_calculating_df[c_calculating_df['Vaccine'].str.contains('-')]

# Calculate investment values for combination vaccines with missing values
for index, row in combination_vaccines[combination_vaccines['Investment'].isnull()].iterrows():
    vaccines = row['Vaccine'].split('-')
    investment_sum = c_calculating_df[c_calculating_df['Vaccine'].isin(vaccines)]['Investment'].sum()
    c_calculating_df.loc[index, 'Investment'] = investment_sum

# Calculate investment values for singular vaccines with missing values
singular_vaccines = c_calculating_df[~c_calculating_df['Vaccine'].str.contains('-')]
singular_vaccines_with_missing_values = singular_vaccines[singular_vaccines['Investment'].isnull()].index
average_investment = c_calculating_df.loc[singular_vaccines.index, 'Investment'].mean()
c_calculating_df.loc[singular_vaccines_with_missing_values, 'Investment'] = average_investment

c_calculating_df.dropna(subset=['Company'], inplace=True)

c_calculating_df['Investment'] = c_calculating_df['Investment'] * precresult
c_calculating_df['Investment'] = .05 * c_calculating_df['Investment'] / (1 - (1 + .05) ** - 20)
c_calculating_df['Capacity, Doses (Total)'] = c_calculating_df['Capacity, Doses (Total)'] * precresult

# Create a dictionary to map vaccine names to indices
vaccine_indices = {vaccine_name: index for index, vaccine_name in vaccine_names.items()}

# Create a dictionary to map producer names to codes
producer_codes = {producer: code for code, producer in paramk.items()}

# Iterate over each row in the c_calculating_df DataFrame
for index, row in c_calculating_df.iterrows():
    # Get the vaccine name and producer name from the row
    producer_name = row['Company']
    
    # Replace the producer name with its corresponding code if available, otherwise keep the original name
    producer_code = next((code for code, name in producer_codes.items() if name == producer_name), producer_name)
    c_calculating_df.at[index, 'Company'] = producer_code

c_calculating_df.rename(columns={'Investment': 'C', 'Capacity, Doses (Total)': 'S', 'Vaccine': 'name', 'Company': 'k'}, inplace=True)
c_calculating_df = c_calculating_df.sort_values('name')
# Print the updated c_calculating_df
print(c_calculating_df)
excel_file = openpyxl.load_workbook(file_path)
sheet_name = "skcname"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    c_calculating_df.to_excel(writer, sheet_name=sheet_name, index=False)

#############################################################################################################

    ### Reservation Prices for the Countries and the Markets ###

#############################################################################################################

import itertools

# Filter the selected_dataframe by vaccine_bundles
vaccine_filter_df = selected_dataframe[selected_dataframe['Vaccine'].isin(vaccine_bundles)]

# Convert 'Relative_weights' and 'avg. price' columns to numeric types
ResPriceFrame['Relative_weight'] = pd.to_numeric(ResPriceFrame['Relative_weight'], errors='coerce')
vaccine_filter_df['Avg. Price'] = pd.to_numeric(vaccine_filter_df['Avg. Price'], errors='coerce')

# Drop rows with non-numeric values
ResPriceFrame = ResPriceFrame.dropna(subset=['Relative_weight'])
vaccine_filter_df = vaccine_filter_df.dropna(subset=['Avg. Price'])

# Create a list of all combinations of weights and prices
combinations = list(itertools.product(ResPriceFrame['Relative_weight'], vaccine_filter_df['Avg. Price']))

# Calculate the products of each combination (skipping non-numeric values)
products = [weight * price for weight, price in combinations if pd.notnull(weight) and pd.notnull(price)]

# Reshape the products into a matrix
country_names = ResPriceFrame.index
vaccine_names = vaccine_filter_df['Vaccine']
num_countries = len(country_names)
num_vaccines = len(vaccine_names)

product_matrix = pd.DataFrame(np.array(products).reshape(num_countries, num_vaccines),
                              index=range(1, num_countries + 1),
                              columns=range(1, num_vaccines + 1))

# Add a unique column for the index
product_matrix.insert(0, "Index", range(1, num_countries + 1))

# Rename the columns without "Vaccine" prefix
product_matrix.columns = product_matrix.columns.map(str)

# Print the result matrix
print(product_matrix)


excel_file = openpyxl.load_workbook(file_path)
sheet_name = "res_price"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    product_matrix.to_excel(writer, sheet_name=sheet_name, index=False)

# Assuming product_matrix is your existing dataframe

# Get the column names from the product_matrix
column_names = list(product_matrix.columns)

# Add a new column for the market
product_matrix['Market'] = 0

# Iterate over each market and its assigned countries
for market, selected_user in markets.items():
    # Find the row indices corresponding to the selected countries in the product_matrix
    rows = [country_codes[country] for country in selected_user]

    # Update the 'Market' column for the selected rows with the market number
    product_matrix.loc[rows, 'Market'] = market

print(product_matrix)

# Group the product_matrix by the 'Market' column
grouped_product_matrix = product_matrix.groupby('Market')

# Calculate the average for each column within each market
new_market_averages = grouped_product_matrix.mean()

# Reset the index of the market_averages DataFrame
new_market_averages.reset_index(inplace=True)

# Remove the index column from the market_averages DataFrame
new_market_averages = new_market_averages.drop('Index', axis=1)

# Print the market_averages DataFrame
print(new_market_averages)

# Group the product_matrix by the 'Market' column
grouped_product_matrix = product_matrix.groupby('Market')

# Calculate the minimum for each column within each market
market_minima = grouped_product_matrix.min()

# Reset the index of the market_minima DataFrame
market_minima.reset_index(inplace=True)

# Remove the index column from the market_minima DataFrame
market_minima = market_minima.drop('Index', axis=1)

# Print the market_minima DataFrame
print(market_minima)

###############################################################################################################################################################

### Doses ###

###############################################################################################################################################################

# Extract unique antigens from the 'Vaccine' column, overwrite the initial antigen set
antigens = set()
for vaccine in selected_dataframe['Vaccine']:
    vaccines = vaccine.split('-')
    for v in vaccines:
        antigens.add(v.strip())
sorted_antigens = sorted(antigens)

# Filter out antigens not present in the updated set
filtered_antigen_doses = {antigen: doses for antigen, doses in antigen_doses.items() if antigen in antigens}

# Create an empty DataFrame with the antigens as index
ad_df = pd.DataFrame(index=filtered_antigen_doses.keys(), columns=markets)

# Fill the DataFrame with the minimum doses for each antigen across markets
for antigen in filtered_antigen_doses:
    ad_df.loc[antigen] = filtered_antigen_doses[antigen]



# Create a new DataFrame with the index from c_calculating_df['name']
vaccine_bundle_df = pd.DataFrame(index=c_calculating_df['name'], columns=markets)

# Fill the new DataFrame with the doses for each vaccine bundle
for bundle in c_calculating_df['name']:
    bundle_vaccines = bundle.split('-')
    bundle_doses = []

    # Iterate over each vaccine in the bundle to get the corresponding doses
    for vaccine in bundle_vaccines:
        if vaccine in ad_df.index:
            doses = ad_df.loc[vaccine]
            bundle_doses.append(doses)

    # Concatenate the doses of the components into a single Series
    if bundle_doses:
        bundle_series = pd.concat(bundle_doses)
        bundle_series = bundle_series.groupby(level=0).min()
        vaccine_bundle_df.loc[bundle] = bundle_series.values

vaccine_bundle_df.sort_index(inplace=True)

# Display the vaccine_bundle_df
print(vaccine_bundle_df)

# Convert antigens to their respective codes in ad_df
ad_df.index = ad_df.index.map(antigen_codes)
# Display the ad_df
print(ad_df)

# Create an empty dataframe with the antigen codes as index
lo = pd.DataFrame(index=ad_df.index, columns=ad_df.columns)

# Iterate over each market and its assigned countries
for market, selected_user in markets.items():
    # Filter the original dataframe for countries in the current market
    market_data_2 = country_param_df[country_param_df['country_name'].isin(selected_user)]

    # Calculate the average births for the market
    average_births = market_data_2['annual_births'].sum()

    # Update the entry in the new dataframe for each antigen code
    for antigen_code in lo.index:
        lo.loc[antigen_code, market] = average_births

# Display the dataframe
print(lo)

excel_file = openpyxl.load_workbook(file_path)
sheet_name = "vacbund"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    writer.book = excel_file
    vaccine_bundle_df.to_excel(writer, sheet_name=sheet_name, index=False)

    #################################################################################

    ### Write output to file ###

    #################################################################################

vaccine_bundle_df.reset_index(drop=True, inplace=True)  # Drop the index and reset it
vaccine_bundle_df.index = range(1, len(vaccine_bundle_df) + 1)  # Reindex with consecutive integers starting from 1 
country_param_df['country_name'] = country_param_df['country_name'].apply(lambda x: f"'{x}'")
product_matrix.drop('Index', axis=1, inplace=True)
c_calculating_df.reset_index(drop=True, inplace=True)
c_calculating_df.index = range(1, len(c_calculating_df) + 1)
new_market_averages.set_index('Market', inplace=True)
new_market_averages = new_market_averages.rename_axis(index=None, columns= None)
market_minima.set_index('Market', inplace=True)
market_minima = market_minima.rename_axis(index = None, columns = None)
product_matrix.drop('Market', axis=1, inplace=True)

def write_output_to_file(output):
    with open("data.dat", "a") as file:
        file.write(output + '\n')

n = f" "
antigen = f"param n_antigens :=" +  str(len(sorted_antigens)) + ";"
nmarkets = f"param n_markets := {num_markets} ;"
ncountries = "param n_countries :=" + str(len(country_names)) +" ;"
bundles = f"param n_bundles :=" + str(len(vaccine_bundles)) + " ;"
producers = f"param n_producers :=" + str(len(setp)) + " ;"
lowuncertainty = "param uncertainty_low := 1.0;"
highuncertainty = "param uncertainty_high := 1.0;"
interestrate = "param interest_rate := 1.05;"
nlinsegments = "param n_linsegments := 5;"
ptyOverIm = "param ptyOverIm:=1e4;"
tenderlength = "param tenderlength:=1;"


write_output_to_file(antigen)
write_output_to_file(nmarkets)
write_output_to_file(ncountries)
write_output_to_file(bundles)
write_output_to_file(producers)
write_output_to_file(n)
write_output_to_file(lowuncertainty)
write_output_to_file(highuncertainty)
write_output_to_file(interestrate)
write_output_to_file(nlinsegments)
write_output_to_file(ptyOverIm)
write_output_to_file(tenderlength)
write_output_to_file(n)

excluded_column = 'Country'
excluded_columm2 = 'Market'

# Open the file in append mode
with open('data.dat', 'a') as file:
    # Write the vaccine bundles along with their codes and antigen values to the file
    for index, bundle in enumerate(vaccine_bundles, start=1):
        antigens = []  # List to store the antigen codes for the current bundle
        
        # Extract the antigens from the bundle and map them to their codes
        for antigen in bundle.split('-'):
            if antigen in antigen_codes:
                antigens.append(str(antigen_codes[antigen]))
        
        # Write the output to the file for each index
        file.write(f"set A1[{index}] := {' '.join(antigens)} ;\n")
    file.write('\n')
    # Print the bundle indices that contain each antigen
    for antigen_code, bundle_indices in antigen_bundle_indices.items():
        antigen = next(key for key, value in antigen_codes.items() if value == antigen_code)
        file.write(f"set B1[{antigen_code}] := {' '.join(map(str, bundle_indices))} ; \n")
    file.write('\n')
    # Print the vaccine indices associated with each producer
    for producer_code, vaccine_indices in producer_vaccine_indices.items():
        producer = next(key for key, value in paramk.items() if value == producer_code)
        file.write(f"set B2[{producer_code}] := {' '.join(map(str, vaccine_indices))} ; \n")
    file.write('\n')

    CkS_column_headings = ' '.join(c_calculating_df.columns)
    file.write('param : ' + CkS_column_headings + ':=')
    file.write('\n')
    c_calculating_df.to_csv(file, index=True, header=False, sep=' ')
    file.write(';\n')
    file.write('param phi := 5;\n')
    file.write('\n')
    file.write('param theta := 1.2;\n')
    file.write('\n')
    file.write('param :' + ' '.join(country_param_df.columns) + ':=')
    file.write('\n')
        # Write the DataFrame data without additional new lines
    country_param_df.to_csv(file, index=True, header=False, sep=' ')
    file.write(';\n')
    file.write('\n')
    file.write('param Rbc (tr) :' + ' '.join(product_matrix.columns) + ':=')
    file.write('\n')
    product_matrix.to_csv(file, index=True, header=False, sep=' ')
    file.write(';\n')
    cheaders = [col for col in marketa.columns if col != excluded_column]
    file.write('param Countries : ' + ' '.join(cheaders) + ' :=')
    file.write('\n')
    marketa.to_csv(file, index=False, header=False, sep=' ')
    file.write(' ;\n')
    aheaders = [col for col in market_averages.columns if col != excluded_columm2 and col != excluded_column]
    file.write('param :' + ' '.join(aheaders) + ' :=')
    file.write('\n')
    market_averages_without_column = market_averages.loc[:, market_averages.columns != excluded_column]
    market_averages_without_column.to_csv(file, index=False, header=False, sep=' ')
    file.write(' ;\n')
    lheaders = [str(col) for col in lo.columns]
    file.write('param lo : ' + ' '.join(lheaders) + ':=')
    file.write('\n')
    lo.to_csv(file, index=True, header=False, sep=' ')
    file.write(' ;\n')
    file.write('param R (tr) : ' + ' '.join(new_market_averages.columns) + ':=')
    file.write('\n')
    new_market_averages.to_csv(file, index=True, header=False, sep=' ')
    file.write(' ;\n')
    file.write('param min_price (tr) : ' + ' '.join(market_minima.columns) + ':=')
    file.write('\n')
    market_minima.to_csv(file, index=True, header=False, sep=' ')
    file.write(' ;\n')
    file.write('param d :' + ' '.join(str(col) for col in ad_df.columns) + ':=')
    file.write('\n')
    ad_df.to_csv(file, index=True, header=False, sep=' ')
    file.write(' ;\n')
    file.write('param D : ' + ' '.join(str(col) for col in vaccine_bundle_df.columns) + ':=')
    file.write('\n')
    vaccine_bundle_df.to_csv(file, index=True, header=False, sep=' ')
    file.write(' ;\n')

file.close()
    
    
    


    
    
    










    

    

